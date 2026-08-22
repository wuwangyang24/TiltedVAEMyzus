"""Generate per-dataset Excel workbooks from BestValLossReporter txt reports.

Each dataset folder (e.g. ``reports/resnet18/mammals/``) is expected to contain
one ``*_best_val_loss_report.txt`` per training run: a Vanilla (SupCon) baseline
and one or more SupConSoftPos runs with different positive-weight temperatures τ.

For every dataset folder found, this script writes a workbook with four
worksheets (Order, Family, Genus, Specific Epithet). Each worksheet tabulates
kNN@1 / kNN@5 / LinProbe@1 / LinProbe@5 for the fixed τ rows plus the Vanilla
baseline, showing each non-Vanilla score together with its change (Δ) versus
Vanilla. All scores and Δ are multiplied by 100; τ values are left unchanged.

Usage
-----
    python generate_reports.py --reports_dir results/reports
    python generate_reports.py --reports_dir results/reports --output_dir excel
"""

import argparse
import os
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ── Fixed layout (from the task specification) ────────────────────────────────
# τ rows are discovered dynamically from the reports (see build_workbook);
# every τ found is included, sorted ascending, with Vanilla last.

# Worksheet name -> the eval-prefix suffix to look up in the parsed report
# (matches val_train_order, val_test_family, val_test_genus,
# val_test_specific_epithet regardless of the train/test prefix).
WORKSHEETS = [
    ("Order", "order"),
    ("Family", "family"),
    ("Genus", "genus"),
    ("Specific Epithet", "specific_epithet"),
]

# Display column header -> metric key suffix used in the report rows.
METRIC_COLUMNS = [
    ("kNN@1", "knn_top1"),
    ("kNN@5", "knn_top5"),
    ("LinProbe@1", "linprobe_top1"),
    ("LinProbe@5", "linprobe_top5"),
]

MINUS = "\u2212"  # unicode minus, matches the "(−18.26)" example


def _to_float(token: str):
    token = token.strip()
    if token in ("", "-", "\u2013", "\u2014"):
        return None
    try:
        return float(token)
    except ValueError:
        return None


def parse_report(path: str):
    """Return ``(kind, metrics)`` for a single report file.

    ``kind`` is either ``"vanilla"`` or a float τ. ``metrics`` maps an eval
    prefix (e.g. ``val_test_family``) to a dict of metric-key -> value.
    """
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()

    fname = os.path.basename(path).lower()

    # Identify the run: Vanilla baseline vs a SupConSoftPos τ run.
    kind = None
    if "vanilla" in fname:
        kind = "vanilla"
    if kind is None:
        m = re.search(r"positive-weight\s+tau:\s*([0-9]*\.?[0-9]+)", text, re.I)
        if m:
            kind = float(m.group(1))
    if kind is None:
        m = re.search(r"tau([0-9]*\.?[0-9]+)", fname)
        if m:
            kind = float(m.group(1))
    if kind is None and re.search(r"positive-weight\s+tau:\s*n/?a", text, re.I):
        kind = "vanilla"

    # Extract the metric table rows: "prefix | v | v | v | v".
    metrics = {}
    for line in text.splitlines():
        if "|" not in line:
            continue
        cells = [c.strip() for c in line.split("|")]
        prefix = cells[0]
        if not prefix.startswith(("val_train_", "val_test_")):
            continue
        if len(cells) < 5:
            continue
        values = [_to_float(c) for c in cells[1:5]]
        metrics[prefix] = {
            "knn_top1": values[0],
            "knn_top5": values[1],
            "linprobe_top1": values[2],
            "linprobe_top5": values[3],
        }
    return kind, metrics


def collect_dataset(report_paths):
    """Parse all reports for one dataset into ``{row_label: {prefix: metrics}}``.

    ``row_label`` is one of the fixed τ labels or ``"Vanilla"``.
    """
    by_kind = {}
    for path in report_paths:
        kind, metrics = parse_report(path)
        if kind is None:
            print(f"  [skip] could not identify run type: {os.path.basename(path)}")
            continue
        by_kind[kind] = metrics
    return by_kind


def _find_metrics(metrics: dict, suffix: str):
    """Return the metric dict whose eval prefix ends with ``_<suffix>``."""
    for prefix, values in metrics.items():
        if prefix.endswith(f"_{suffix}"):
            return values
    return None


def build_worksheet(ws, suffix: str, by_kind: dict, tau_rows):
    center = Alignment(horizontal="center", vertical="center")

    header = ["\u03c4 / Method"] + [c[0] for c in METRIC_COLUMNS]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = center

    # Assemble the raw scores (×100) for each row and metric.
    row_labels = [f"\u03c4 = {tau:g}" for tau in tau_rows] + ["Vanilla"]
    kinds = list(tau_rows) + ["vanilla"]

    vanilla_metrics = _find_metrics(by_kind.get("vanilla", {}), suffix) or {}
    baseline = {
        mkey: (vanilla_metrics.get(mkey) * 100 if vanilla_metrics.get(mkey) is not None else None)
        for _, mkey in METRIC_COLUMNS
    }

    # scores[row_index][metric_index] = float score (×100) or None
    scores = []
    for kind in kinds:
        m = _find_metrics(by_kind.get(kind, {}), suffix) or {}
        scores.append([
            (m.get(mkey) * 100 if m.get(mkey) is not None else None)
            for _, mkey in METRIC_COLUMNS
        ])

    # Column maxima (by score, ignoring Δ) for bolding.
    col_max = []
    for j in range(len(METRIC_COLUMNS)):
        vals = [scores[i][j] for i in range(len(kinds)) if scores[i][j] is not None]
        col_max.append(max(vals) if vals else None)

    # Write rows.
    for i, (label, kind) in enumerate(zip(row_labels, kinds)):
        row_cells = [label]
        for j, (_, mkey) in enumerate(METRIC_COLUMNS):
            score = scores[i][j]
            if score is None:
                row_cells.append("-")
                continue
            if kind == "vanilla":
                row_cells.append(f"{score:.2f}")
            else:
                base = baseline.get(mkey)
                if base is None:
                    row_cells.append(f"{score:.2f}")
                else:
                    delta = score - base
                    delta_str = f"{delta:+.2f}".replace("-", MINUS)
                    row_cells.append(f"{score:.2f} ({delta_str})")
        ws.append(row_cells)

        excel_row = i + 2  # +1 header, +1 for 1-based indexing
        for j in range(len(header)):
            cell = ws.cell(row=excel_row, column=j + 1)
            cell.alignment = center
            # Bold winning metric cells (column j-1 in scores; col 0 is label).
            if j >= 1:
                score = scores[i][j - 1]
                cmax = col_max[j - 1]
                if score is not None and cmax is not None and abs(score - cmax) < 1e-9:
                    cell.font = Font(bold=True)

    # Freeze header row.
    ws.freeze_panes = "A2"

    # Auto-fit column widths.
    for col in range(1, len(header) + 1):
        letter = get_column_letter(col)
        width = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = width + 4


def build_workbook(by_kind: dict, out_path: str):
    # τ rows are discovered from the reports (sorted ascending); Vanilla is
    # always rendered last as the baseline row.
    tau_rows = sorted(k for k in by_kind if k != "vanilla")
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, suffix in WORKSHEETS:
        ws = wb.create_sheet(title=sheet_name)
        build_worksheet(ws, suffix, by_kind, tau_rows)
    wb.save(out_path)


def find_dataset_dirs(reports_dir: str):
    """Yield (dir_path, [report_files]) for each folder with report txts."""
    groups = defaultdict(list)
    for root, _dirs, files in os.walk(reports_dir):
        for name in files:
            if name.endswith("_best_val_loss_report.txt") or (
                name.endswith(".txt") and "report" in name.lower()
            ):
                groups[root].append(os.path.join(root, name))
    return groups


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--reports_dir", default="reports",
                   help="Root folder containing <model>/<dataset>/*_report.txt files.")
    p.add_argument("--output_dir", default=None,
                   help="Where to write the .xlsx files. Defaults to each dataset folder.")
    args = p.parse_args()

    groups = find_dataset_dirs(args.reports_dir)
    if not groups:
        print(f"No report files found under: {args.reports_dir}")
        return

    if args.output_dir:
        os.makedirs(args.output_dir, exist_ok=True)

    for folder, report_paths in sorted(groups.items()):
        by_kind = collect_dataset(report_paths)
        if not by_kind:
            print(f"[skip] no parseable reports in {folder}")
            continue

        rel = os.path.relpath(folder, args.reports_dir)
        name = rel.replace(os.sep, "_") if rel != "." else os.path.basename(folder.rstrip(os.sep))
        out_dir = args.output_dir or folder
        out_path = os.path.join(out_dir, f"{name}_metrics.xlsx")

        build_workbook(by_kind, out_path)
        found = ", ".join(
            ["Vanilla" if k == "vanilla" else f"\u03c4={k:g}" for k in sorted(
                by_kind, key=lambda x: (x == "vanilla", x))]
        )
        print(f"[ok] {out_path}  (runs: {found})")


if __name__ == "__main__":
    main()
